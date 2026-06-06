"""
PureCuts Product Upload Script
================================
Uploads all products from 'Product Details.xlsx' to Firebase Firestore.

Usage:
    1. Place your Firebase service account key at:
       C:\Users\manep\purecuts\scripts\serviceAccountKey.json
    2. Run: python upload_products.py

Requirements:
    pip install firebase-admin openpyxl
"""

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore
import uuid
import os
import re

# ─── CONFIG ──────────────────────────────────────────────────────────────────

EXCEL_FILE = r"C:\Users\manep\purecuts\Product Details.xlsx"
SERVICE_ACCOUNT_KEY = r"C:\Users\manep\purecuts\scripts\serviceAccountKey.json"
COLLECTION_NAME = "products"
BATCH_SIZE = 400  # Firestore max is 500 per batch

# ─── INIT FIREBASE ───────────────────────────────────────────────────────────

if not os.path.exists(SERVICE_ACCOUNT_KEY):
    print(f"ERROR: Service account key not found at:\n  {SERVICE_ACCOUNT_KEY}")
    print("\nTo get it:")
    print("  1. Go to Firebase Console → Project Settings → Service Accounts")
    print("  2. Click 'Generate new private key'")
    print("  3. Save the downloaded JSON as 'serviceAccountKey.json' in the scripts/ folder")
    exit(1)

cred = credentials.Certificate(SERVICE_ACCOUNT_KEY)
firebase_admin.initialize_app(cred)
db = firestore.client()

# ─── READ EXCEL ───────────────────────────────────────────────────────────────

print(f"Reading Excel file: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print(f"Columns: {headers}")
print(f"Total rows: {ws.max_row - 1} products\n")

# Map column names to indices (0-based)
def col(name):
    return headers.index(name)

# ─── PARSE & UPLOAD ───────────────────────────────────────────────────────────

def clean_tags(raw):
    """Convert tag string like 'Creams, bodycare' to list ['Creams', 'bodycare']"""
    if not raw or raw == 'NULL':
        return []
    return [t.strip() for t in re.split(r'[,;]+', str(raw)) if t.strip()]


def make_product_id(row_num, name):
    """Create a stable, readable product ID"""
    slug = re.sub(r'[^a-z0-9]+', '_', str(name).lower())[:30].strip('_')
    return f"prod_{row_num:04d}_{slug}"


products = []
skipped = 0

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    try:
        name = str(row[col('Name')]).strip() if row[col('Name')] else None
        if not name or name == 'None':
            skipped += 1
            continue

        mrp = float(row[col('MRP')]) if row[col('MRP')] else 0.0
        selling_price = float(row[col('Selling price')]) if row[col('Selling price')] else mrp
        image_url = str(row[col('Image URL')]).strip() if row[col('Image URL')] else ''
        weight = str(row[col('Weight ')]).strip() if row[col('Weight ')] else ''
        tags = clean_tags(row[col('Tags')])

        product_id = make_product_id(i, name)

        product = {
            'id': product_id,
            'name': name,
            'brand': str(row[col('Brand')]).strip() if row[col('Brand')] else '',
            'category': str(row[col('Category')]).strip() if row[col('Category')] else '',
            'subcategory': str(row[col('Subcategory')]).strip() if row[col('Subcategory')] else '',
            'price': selling_price,
            'originalPrice': mrp,
            'imageUrl': image_url,
            'size': weight if weight and weight.lower() != 'null' else '',
            'tags': tags,
            'stock': 50,          # default stock — update via dashboard later
            'rating': 4.0,        # default rating
            'reviews': 0,
            'isPopular': False,
            'isRecommended': False,
            'deliveryTime': '2-3 days',
            'description': '',
            'createdAt': firestore.SERVER_TIMESTAMP,
        }

        products.append((product_id, product))

    except Exception as e:
        print(f"  Row {i} skipped — error: {e}")
        skipped += 1

print(f"Parsed {len(products)} products ({skipped} skipped)\n")

# ─── BATCH WRITE TO FIRESTORE ─────────────────────────────────────────────────

total_uploaded = 0
batch_num = 0

for batch_start in range(0, len(products), BATCH_SIZE):
    batch_num += 1
    batch_slice = products[batch_start:batch_start + BATCH_SIZE]
    batch = db.batch()

    for product_id, product in batch_slice:
        doc_ref = db.collection(COLLECTION_NAME).document(product_id)
        batch.set(doc_ref, product)

    batch.commit()
    total_uploaded += len(batch_slice)
    print(f"Batch {batch_num}: uploaded {len(batch_slice)} products  (total: {total_uploaded}/{len(products)})")

print(f"\n✅ Done! {total_uploaded} products uploaded to Firestore collection '{COLLECTION_NAME}'")
print("\nNext steps:")
print("  1. Open Firebase Console → Firestore → 'products' collection to verify")
print("  2. The Flutter app will now load real products from Firestore")
print("  3. Use the admin dashboard to edit/update products going forward")
